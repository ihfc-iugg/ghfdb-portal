"""Round-trip regression tests for GHFDB import -> export (SC-001)."""

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest


def _norm_empty(value):
    """Normalize None/empty-like values to empty string for stable comparisons."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _assert_float_close(actual, expected, tol=1e-9):
    """Assert actual ~= expected for numeric export cells."""
    if expected == "":
        assert _norm_empty(actual) == ""
        return
    assert abs(float(actual) - float(expected)) < tol


@pytest.mark.django_db
def test_roundtrip_import_then_export_preserves_values(dataset):
    """
    SC-001:
    1) Import fixture XLSX using parent + child resources
    2) Export using GHFDBExportResource
    3) Verify text/vocabulary equality and numeric closeness
    """
    from project.ghfdb.models import GHFDB
    from project.ghfdb.resources import (
        GHFDBChildImportResource,
        GHFDBExportResource,
        GHFDBImportFormat,
        GHFDBParentImportResource,
    )

    fixture_path = Path(__file__).resolve().parents[1] / "fixtures" / "sample_ghfdb.xlsx"
    fixture_bytes = fixture_path.read_bytes()

    # Parse twice because parent import mutates its Dataset during deduplication.
    fmt = GHFDBImportFormat()
    ds_parent = fmt.create_dataset(fixture_bytes)
    ds_child = fmt.create_dataset(fixture_bytes)

    parent_resource = GHFDBParentImportResource()
    parent_result = parent_resource.import_data(
        ds_parent,
        dry_run=False,
        raise_errors=True,
        fairdm_dataset=dataset,
    )
    assert not parent_result.has_errors(), parent_result.invalid_rows

    child_resource = GHFDBChildImportResource()
    child_result = child_resource.import_data(
        ds_child,
        dry_run=False,
        raise_errors=True,
        fairdm_dataset=dataset,
    )
    assert not child_result.has_errors(), child_result.invalid_rows

    export_resource = GHFDBExportResource()
    export_qs = GHFDB.objects.for_export().order_by("ghfdb_id")
    exported_rows = list(export_resource.export(export_qs).dict)

    assert len(exported_rows) == 3

    by_comment = {row["c_comment"]: row for row in exported_rows}

    expected = {
        "child-corrections": {
            "text": {
                "name": "Roundtrip Site",
                "environment": "onshore_continental",
                "p_comment": "parent comment",
                "explo_purpose": "Hydrocarbon",
                "q_method": "Bullard method",
                "corr_is_flag": "present_corrected",
                "corr_t_flag": "present_not_corrected",
                "corr_s_flag": "present_not_significant",
                "corr_e_flag": "not_recognized",
                "corr_topo_flag": "considered_p",
                "corr_pal_flag": "considered_t",
                "corr_sur_flag": "considered_pt",
                "corr_conv_flag": "not_considered",
                "corr_hr_flag": "present_not_significant",
                "expedition": "Expedition A",
            },
            "numeric": {
                "q": 70,
                "q_uncertainty": 5,
                "qc": 68,
                "qc_uncertainty": 4,
                "q_top": 0,
                "q_bottom": 200,
                "tc_mean": "",
                "t_grad_mean": "",
            },
        },
        "child-gradient-conductivity": {
            "text": {
                "name": "Roundtrip Site",
                "environment": "onshore_continental",
                "explo_purpose": "Hydrocarbon",
                "q_method": "Interval method",
                "corr_is_flag": "-",
                "corr_t_flag": "-",
                "corr_s_flag": "-",
                "corr_e_flag": "-",
                "corr_topo_flag": "-",
                "corr_pal_flag": "-",
                "corr_sur_flag": "-",
                "corr_conv_flag": "-",
                "corr_hr_flag": "-",
                "expedition": "Expedition B",
                "t_method_top": "BHT",
                "t_method_bottom": "BLK",
                "t_corr_top": "AAPG correction",
                "t_corr_bottom": "Horner plot",
                "tc_source": "Core samples",
                "tc_location": "Actual heat-flow location",
                "tc_method": "Estimation - from lithology and literature",
                "tc_saturation": "Dry measured",
                "tc_pT_conditions": "Actual in-situ (pT) conditions",
                "tc_pT_function": "Other",
                "tc_strategy": "Characterize formation conductivities",
            },
            "numeric": {
                "q": 70,
                "q_uncertainty": 5,
                "qc": 66,
                "qc_uncertainty": 3,
                "q_top": 200,
                "q_bottom": 500,
                "probe_penetration": 2,
                "probe_length": 3,
                "probe_tilt": 5,
                "water_temperature": 3,
                "t_grad_mean": 25,
                "t_grad_uncertainty": 1,
                "t_grad_mean_cor": 24,
                "t_grad_uncertainty_cor": 0.8,
                "t_shutin_top": 24,
                "t_shutin_bottom": 36,
                "t_number": 8,
                "tc_mean": 2.5,
                "tc_uncertainty": 0.2,
                "tc_number": 5,
            },
        },
        "child-minimal": {
            "text": {
                "name": "Roundtrip Site",
                "environment": "onshore_continental",
                "explo_purpose": "Hydrocarbon",
                "q_method": "Boot-strapping method",
                "corr_is_flag": "-",
                "corr_t_flag": "-",
                "corr_s_flag": "-",
                "corr_e_flag": "-",
                "corr_topo_flag": "-",
                "corr_pal_flag": "-",
                "corr_sur_flag": "-",
                "corr_conv_flag": "-",
                "corr_hr_flag": "-",
            },
            "numeric": {
                "q": 70,
                "q_uncertainty": 5,
                "qc": 64,
                "qc_uncertainty": 2,
                "q_top": 500,
                "q_bottom": 800,
                "tc_mean": "",
                "t_grad_mean": "",
            },
        },
    }

    for comment, expected_row in expected.items():
        actual = by_comment[comment]

        for col, exp in expected_row["text"].items():
            assert _norm_empty(actual[col]) == _norm_empty(exp), f"Mismatch in '{col}' for row '{comment}'"

        for col, exp in expected_row["numeric"].items():
            _assert_float_close(actual[col], exp)


def _build_simple_xlsx_from_official(official_xlsx_bytes: bytes) -> bytes:
    """Convert an official GHFDB XLSX to simple layout by removing rows 7 and 8.

    The official template has:
      Row 6: headers, Row 7: units, Row 8: ranges, Row 9+: data

    The simple layout has:
      Row 6: headers, Row 7+: data

    This function reads the official XLSX, drops rows 7-8, and re-writes
    a new XLSX in the simple layout.

    Args:
        official_xlsx_bytes: Raw bytes of the official-layout XLSX.

    Returns:
        Raw bytes of the equivalent simple-layout XLSX.
    """
    wb_in = openpyxl.load_workbook(BytesIO(official_xlsx_bytes), read_only=True, data_only=True)
    ws_in = wb_in["data list"]

    # Read all rows (1-indexed); official has headers at row 6, unit at 7, range at 8
    all_rows = list(ws_in.iter_rows(values_only=True))
    # Rows 1-6 (0-indexed 0-5) stay; rows 7-8 (0-indexed 6-7) are dropped; data is 8+ (0-indexed)
    kept_rows = all_rows[:6] + all_rows[8:]  # keep metadata+header, skip units/ranges, keep data
    wb_in.close()

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "data list"

    for row_idx, row_values in enumerate(kept_rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            ws_out.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb_out.save(buf)
    return buf.getvalue()


@pytest.mark.django_db
def test_roundtrip_simple_format_import_then_export_preserves_values(dataset):
    """
    SC-001 (simple template variant):
    1) Convert fixture XLSX to simple layout (remove unit/range rows 7-8)
    2) Import using GHFDBSimpleImportFormat (parent + child resources)
    3) Export using GHFDBExportResource
    4) Verify identical text/vocabulary equality and numeric closeness as official format test
    """
    from project.ghfdb.models import GHFDB
    from project.ghfdb.resources import (
        GHFDBChildImportResource,
        GHFDBExportResource,
        GHFDBParentImportResource,
        GHFDBSimpleImportFormat,
    )

    fixture_path = Path(__file__).resolve().parents[1] / "fixtures" / "sample_ghfdb.xlsx"
    official_bytes = fixture_path.read_bytes()
    simple_bytes = _build_simple_xlsx_from_official(official_bytes)

    fmt = GHFDBSimpleImportFormat()
    ds_parent = fmt.create_dataset(simple_bytes)
    ds_child = fmt.create_dataset(simple_bytes)

    parent_resource = GHFDBParentImportResource()
    parent_result = parent_resource.import_data(
        ds_parent,
        dry_run=False,
        raise_errors=True,
        fairdm_dataset=dataset,
    )
    assert not parent_result.has_errors(), parent_result.invalid_rows

    child_resource = GHFDBChildImportResource()
    child_result = child_resource.import_data(
        ds_child,
        dry_run=False,
        raise_errors=True,
        fairdm_dataset=dataset,
    )
    assert not child_result.has_errors(), child_result.invalid_rows

    export_resource = GHFDBExportResource()
    export_qs = GHFDB.objects.for_export().order_by("ghfdb_id")
    exported_rows = list(export_resource.export(export_qs).dict)

    assert len(exported_rows) == 3

    by_comment = {row["c_comment"]: row for row in exported_rows}

    expected = {
        "child-corrections": {
            "text": {
                "name": "Roundtrip Site",
                "environment": "onshore_continental",
                "p_comment": "parent comment",
                "explo_purpose": "Hydrocarbon",
                "q_method": "Bullard method",
                "corr_is_flag": "present_corrected",
                "corr_t_flag": "present_not_corrected",
                "corr_s_flag": "present_not_significant",
                "corr_e_flag": "not_recognized",
                "corr_topo_flag": "considered_p",
                "corr_pal_flag": "considered_t",
                "corr_sur_flag": "considered_pt",
                "corr_conv_flag": "not_considered",
                "corr_hr_flag": "present_not_significant",
                "expedition": "Expedition A",
            },
            "numeric": {
                "q": 70,
                "q_uncertainty": 5,
                "qc": 68,
                "qc_uncertainty": 4,
                "q_top": 0,
                "q_bottom": 200,
                "tc_mean": "",
                "t_grad_mean": "",
            },
        },
        "child-gradient-conductivity": {
            "text": {
                "name": "Roundtrip Site",
                "environment": "onshore_continental",
                "explo_purpose": "Hydrocarbon",
                "q_method": "Interval method",
            },
            "numeric": {
                "q": 70,
                "q_uncertainty": 5,
                "qc": 66,
                "qc_uncertainty": 3,
                "q_top": 200,
                "q_bottom": 500,
            },
        },
        "child-minimal": {
            "text": {
                "name": "Roundtrip Site",
                "environment": "onshore_continental",
                "explo_purpose": "Hydrocarbon",
                "q_method": "Boot-strapping method",
            },
            "numeric": {
                "q": 70,
                "q_uncertainty": 5,
                "qc": 64,
                "qc_uncertainty": 2,
                "q_top": 500,
                "q_bottom": 800,
                "tc_mean": "",
                "t_grad_mean": "",
            },
        },
    }

    for comment, expected_row in expected.items():
        actual = by_comment[comment]

        for col, exp in expected_row["text"].items():
            assert _norm_empty(actual[col]) == _norm_empty(exp), f"[simple] Mismatch in '{col}' for row '{comment}'"

        for col, exp in expected_row["numeric"].items():
            _assert_float_close(actual[col], exp)
