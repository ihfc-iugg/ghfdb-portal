"""
Tests for GHFDBParentImportResource.

Covers:
- Upsert on ghfdb_id (re-import updates, does not duplicate)
- before_import() deduplication keeps first occurrence of each ID_parent
- 18 parent columns mapped to correct fields
- ParentHeatFlow.sample FK (to HeatFlowSite) created with correct Point location
- explo_purpose M2M set
- Staff-only access control (anonymous admin import URL -> 302)
"""

import pytest
import tablib
from django.contrib.auth import get_user_model

User = get_user_model()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

PARENT_ROW = {
    "ID_parent": "1",
    "q": "70.0",
    "q_uncertainty": "5.0",
    "name": "Test Site Alpha",
    "lat_NS": "48.0",
    "long_EW": "11.0",
    "elevation": "500",
    "environment": "Onshore (continental)",
    "p_comment": "test comment",
    "corr_HP_flag": "No",
    "total_depth_MD": "1000",
    "total_depth_TVD": "900",
    "explo_method": "",
    "explo_purpose": "Hydrocarbon",
    "Country": "Germany",
    "Region": "Bavaria",
    "Continent": "Europe",
    "Domain": "",
}


def make_dataset(*rows):
    """Build a tablib Dataset from dicts with matching headers."""
    headers = list(rows[0].keys())
    ds = tablib.Dataset(headers=headers)
    for row in rows:
        ds.append([row[h] for h in headers])
    return ds


# ---------------------------------------------------------------------------
# T029 Tests
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestGHFDBParentImportResourceImport:
    """T029 — GHFDBParentImportResource end-to-end import tests."""

    def test_import_creates_parent_and_site(self, dataset):
        """Importing a row creates one ParentHeatFlow and one HeatFlowSite."""
        from heat_flow.models import HeatFlowSite, ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        ds = make_dataset(PARENT_ROW)
        result = resource.import_data(ds, dry_run=False, raise_errors=False)

        assert not result.has_errors(), result.invalid_rows
        assert ParentHeatFlow.objects.filter(ghfdb_id=1).exists()
        assert HeatFlowSite.objects.filter(name="Test Site Alpha").exists()

    def test_upsert_on_ghfdb_id_does_not_duplicate(self, dataset):
        """Re-importing the same ID_parent updates, does not create a second record."""
        from heat_flow.models import ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        ds = make_dataset(PARENT_ROW)
        resource.import_data(ds, dry_run=False, raise_errors=False)
        resource.import_data(ds, dry_run=False, raise_errors=False)

        assert ParentHeatFlow.objects.filter(ghfdb_id=1).count() == 1

    def test_upsert_updates_existing_record(self, dataset):
        """Re-importing with changed q value updates the existing ParentHeatFlow."""
        from heat_flow.models import ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        ds1 = make_dataset(PARENT_ROW)
        resource.import_data(ds1, dry_run=False, raise_errors=False)

        updated = dict(PARENT_ROW)
        updated["q"] = "80.0"
        ds2 = make_dataset(updated)
        resource.import_data(ds2, dry_run=False, raise_errors=False)

        parent = ParentHeatFlow.objects.get(ghfdb_id=1)
        assert float(parent.value.magnitude) == pytest.approx(80.0)

    def test_site_has_correct_point_location(self, dataset):
        """Imported HeatFlowSite has a Point at (long_EW, lat_NS)."""
        from heat_flow.models import HeatFlowSite

        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        ds = make_dataset(PARENT_ROW)
        resource.import_data(ds, dry_run=False, raise_errors=False)

        site = HeatFlowSite.objects.get(name="Test Site Alpha")
        assert site.location is not None
        assert float(site.location.x) == pytest.approx(11.0)
        assert float(site.location.y) == pytest.approx(48.0)

    def test_explo_purpose_m2m_set(self, dataset):
        """Imported HeatFlowSite has explo_purpose M2M populated."""
        from heat_flow.models import HeatFlowSite

        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        ds = make_dataset(PARENT_ROW)
        resource.import_data(ds, dry_run=False, raise_errors=False)

        site = HeatFlowSite.objects.get(name="Test Site Alpha")
        assert site.explo_purpose.count() >= 1

    def test_all_18_parent_columns_accepted(self, dataset):
        """Resource accepts all 18 PARENT_COLUMNS without field errors."""
        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        ds = make_dataset(PARENT_ROW)
        result = resource.import_data(ds, dry_run=True, raise_errors=False)
        assert not result.has_errors(), result.invalid_rows


@pytest.mark.django_db
class TestGHFDBParentBeforeImportDedup:
    """T029 — before_import() deduplication."""

    def test_dedup_keeps_first_row_per_id_parent(self, dataset):
        """before_import() keeps only the first row for each ID_parent."""
        from heat_flow.models import ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        row1 = dict(PARENT_ROW)
        row2 = dict(PARENT_ROW)
        row2["q"] = "99.9"  # different value but same ID_parent

        ds = make_dataset(row1, row2)
        resource.import_data(ds, dry_run=False, raise_errors=False)

        assert ParentHeatFlow.objects.filter(ghfdb_id=1).count() == 1
        parent = ParentHeatFlow.objects.get(ghfdb_id=1)
        # First row's value should be used
        assert float(parent.value.magnitude) == pytest.approx(70.0)


@pytest.mark.django_db
class TestGHFDBParentTemplateNoIdRegression:
    """T067/T029 regression coverage for template uploads without ID_parent."""

    def test_no_id_parent_dedup_uses_lat_long_natural_key(self, dataset):
        """Rows without ID_parent deduplicate by lat_NS + long_EW."""
        from heat_flow.models import HeatFlowSite, ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        row1 = dict(PARENT_ROW)
        row1["ID_parent"] = ""
        row2 = dict(PARENT_ROW)
        row2["ID_parent"] = ""
        row2["q"] = "99.9"  # would overwrite only if second row were imported

        resource = GHFDBParentImportResource()
        ds = make_dataset(row1, row2)
        result = resource.import_data(ds, dry_run=False, raise_errors=False)

        assert not result.has_errors(), result.invalid_rows
        assert ParentHeatFlow.objects.count() == 1
        assert HeatFlowSite.objects.count() == 1
        parent = ParentHeatFlow.objects.first()
        assert parent is not None
        assert float(parent.value.magnitude) == pytest.approx(70.0)

    def test_absent_id_parent_header_does_not_raise_header_error(self, dataset):
        """Import succeeds when ID_parent column is entirely absent from headers."""
        from heat_flow.models import HeatFlowSite, ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        row = {k: v for k, v in PARENT_ROW.items() if k != "ID_parent"}
        ds = make_dataset(row)
        assert "ID_parent" not in ds.headers

        result = GHFDBParentImportResource().import_data(
            ds, dry_run=False, raise_errors=False
        )

        assert not result.has_errors(), result.invalid_rows
        assert ParentHeatFlow.objects.count() == 1
        assert HeatFlowSite.objects.count() == 1

    def test_absent_id_parent_header_reimport_upserts(self, dataset):
        """Re-import without ID_parent header updates records rather than duplicating."""
        from heat_flow.models import ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        row = {k: v for k, v in PARENT_ROW.items() if k != "ID_parent"}
        GHFDBParentImportResource().import_data(
            make_dataset(row), dry_run=False, raise_errors=False
        )

        row_update = dict(row)
        row_update["q"] = "88.8"
        result = GHFDBParentImportResource().import_data(
            make_dataset(row_update), dry_run=False, raise_errors=False
        )

        assert not result.has_errors(), result.invalid_rows
        assert ParentHeatFlow.objects.count() == 1
        parent = ParentHeatFlow.objects.first()
        assert parent is not None
        assert float(parent.value.magnitude) == pytest.approx(88.8)

    def test_no_id_parent_reimport_updates_existing_parent(self, dataset):
        """Re-import without ID_parent upserts via lat_NS + long_EW key."""
        from heat_flow.models import ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        row = dict(PARENT_ROW)
        row["ID_parent"] = ""

        resource.import_data(make_dataset(row), dry_run=False, raise_errors=False)

        row_update = dict(row)
        row_update["q"] = "81.5"
        result = resource.import_data(
            make_dataset(row_update),
            dry_run=False,
            raise_errors=False,
        )

        assert not result.has_errors(), result.invalid_rows
        assert ParentHeatFlow.objects.count() == 1
        parent = ParentHeatFlow.objects.first()
        assert parent is not None
        assert float(parent.value.magnitude) == pytest.approx(81.5)


@pytest.mark.django_db
class TestGHFDBParentImportSiteIdentity:
    """T043/T044 — coordinates, not ID_parent, identify the site (FR-004, R1)."""

    def test_import_resolves_to_existing_site_by_coordinates(self, dataset):
        """
        T043 – Importing a row whose coordinates match an existing site
        resolves to that site rather than creating a second one, even
        though the row carries an ID_parent the site has never seen.
        """
        from fairdm.contrib.location.models import Point
        from heat_flow.models import HeatFlowSite, ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        point = Point.objects.create(x=11.0, y=48.0)
        existing_site = HeatFlowSite.objects.create(
            dataset=dataset, name="Pre-existing Site", location=point
        )

        resource = GHFDBParentImportResource()
        ds = make_dataset(PARENT_ROW)  # lat_NS=48.0, long_EW=11.0, ID_parent=1
        result = resource.import_data(ds, dry_run=False, raise_errors=False)

        assert not result.has_errors(), result.invalid_rows
        assert not result.has_validation_errors(), [
            row.error_dict for row in result.invalid_rows
        ]
        assert HeatFlowSite.objects.count() == 1
        parent = ParentHeatFlow.objects.get(ghfdb_id=1)
        assert parent.sample_id == existing_site.pk

    def test_import_with_disagreeing_id_parent_does_not_duplicate_site(self, dataset):
        """
        T044 – A row that carries a site identifier (ID_parent) the site has
        never seen, while its coordinates belong to a different existing
        site, resolves to the site the coordinates identify rather than
        creating a duplicate at that occupied pair.
        """
        from fairdm.contrib.location.models import Point
        from heat_flow.models import HeatFlowSite, ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        # A site already exists at (48.0, 11.0), under a different local_id
        # than the one the incoming row will carry.
        point = Point.objects.create(x=11.0, y=48.0)
        existing_site = HeatFlowSite.objects.create(
            dataset=dataset,
            name="Existing Site",
            local_id="other-id",
            location=point,
        )

        row = dict(PARENT_ROW)
        row["ID_parent"] = "2"  # unseen, and disagrees with local_id "other-id"

        resource = GHFDBParentImportResource()
        result = resource.import_data(
            make_dataset(row), dry_run=False, raise_errors=False
        )

        assert not result.has_errors(), result.invalid_rows
        assert not result.has_validation_errors(), [
            row.error_dict for row in result.invalid_rows
        ]
        assert HeatFlowSite.objects.count() == 1
        parent = ParentHeatFlow.objects.get(ghfdb_id=2)
        assert parent.sample_id == existing_site.pk


@pytest.mark.django_db
class TestGHFDBParentImportRefusesSecondParent:
    """T051 — US-3: a second parent for a site is refused through the import
    path, not only through the model (FR-012, SC-005)."""

    def test_second_parent_at_the_same_site_is_refused_on_import(self, dataset):
        """
        Importing a row whose coordinates resolve to a site that already has
        a ParentHeatFlow is refused: the row is reported as invalid and no
        second ParentHeatFlow is created for that site.
        """
        from heat_flow.models import HeatFlowSite, ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        resource = GHFDBParentImportResource()
        first_result = resource.import_data(
            make_dataset(PARENT_ROW), dry_run=False, raise_errors=False
        )
        assert not first_result.has_errors(), first_result.invalid_rows
        assert ParentHeatFlow.objects.filter(ghfdb_id=1).exists()

        # Same coordinates as PARENT_ROW (lat_NS=48.0, long_EW=11.0), a
        # different, previously unseen ID_parent.
        second_row = dict(PARENT_ROW)
        second_row["ID_parent"] = "2"
        second_row["q"] = "99.0"

        second_result = GHFDBParentImportResource().import_data(
            make_dataset(second_row), dry_run=False, raise_errors=False
        )

        assert second_result.has_validation_errors()
        assert HeatFlowSite.objects.filter(name="Test Site Alpha").count() == 1
        site = HeatFlowSite.objects.get(name="Test Site Alpha")
        assert ParentHeatFlow.objects.filter(sample=site).count() == 1
        assert not ParentHeatFlow.objects.filter(ghfdb_id=2).exists()


class TestGHFDBParentImportResourceAccessControl:
    """T029 — Staff-only access control."""

    def test_anonymous_admin_import_url_redirects(self, client):
        """Anonymous access to admin import URL redirects (302)."""
        from django.urls import reverse

        url = reverse("admin:ghfdb_ghfdbchild_import")
        response = client.get(url)
        assert response.status_code == 302


@pytest.mark.django_db
class TestGHFDBAutoParentKeyRegression:
    """T075 — BUG-005 regression: synthetic key must not pollute ParentHeatFlow fields after no-ID import."""

    def test_no_auto_parent_in_ghfdb_id_after_no_id_import(self, dataset):
        """No ParentHeatFlow.ghfdb_id should contain a synthetic string key after importing a row without ID_parent."""
        from heat_flow.models import ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        row = {k: v for k, v in PARENT_ROW.items() if k != "ID_parent"}
        resource = GHFDBParentImportResource()
        result = resource.import_data(
            make_dataset(row), dry_run=False, raise_errors=False
        )

        assert not result.has_errors(), result.invalid_rows
        for parent in ParentHeatFlow.objects.all():
            # ghfdb_id is a PositiveIntegerField; None is expected for no-ID rows
            assert parent.ghfdb_id is None or isinstance(parent.ghfdb_id, int), (
                f"ParentHeatFlow.ghfdb_id has unexpected type: {parent.ghfdb_id!r}"
            )

    def test_no_auto_parent_in_dry_run_id_parent_column(self, dataset):
        """Dry-run result ID_parent column must show real ghfdb_id or empty, not AUTO_PARENT:."""
        from project.ghfdb.resources import GHFDBParentImportResource

        row = {k: v for k, v in PARENT_ROW.items() if k != "ID_parent"}
        resource = GHFDBParentImportResource()
        result = resource.import_data(
            make_dataset(row), dry_run=True, raise_errors=False
        )

        assert not result.has_errors(), result.invalid_rows
        for row_result in result.rows:
            # diff is a list of (old, new) tuples for each visible field
            for diff_entry in row_result.diff:
                assert "AUTO_PARENT:" not in str(diff_entry), (
                    f"Dry-run diff contains synthetic AUTO_PARENT key: {diff_entry!r}"
                )


class TestGHFDBParentColumnOrderRegression:
    """T077 — BUG-006 regression: get_user_visible_fields() must follow PARENT_COLUMNS order."""

    @pytest.mark.xfail(strict=True, reason=(
        "Half-landed GHFDB canonical column work: the constants moved to the "
        "published spreadsheet casing, but ghfdb_colmeta.json, the resource "
        "field declarations and one manager annotation key did not follow. "
        "Needs debugging, and a decision on the published column vocabulary, "
        "before it can pass. See issue #122."
    ))
    def test_get_user_visible_fields_follows_parent_columns_order(self):
        """GHFDBParentImportResource.get_user_visible_fields() returns fields in PARENT_COLUMNS order."""
        from project.ghfdb.resources import GHFDBParentImportResource
        from project.ghfdb.resources.formats import PARENT_COLUMNS

        resource = GHFDBParentImportResource()
        visible_fields = resource.get_user_visible_fields()
        col_names = [f.column_name for f in visible_fields]

        # Extract only those column_names that appear in PARENT_COLUMNS
        ordered_present = [c for c in col_names if c in PARENT_COLUMNS]
        # They must appear in the exact order defined by PARENT_COLUMNS
        expected_order = [c for c in PARENT_COLUMNS if c in col_names]
        assert ordered_present == expected_order, (
            f"Column order mismatch.\nGot: {ordered_present}\nExpected: {expected_order}"
        )


# ---------------------------------------------------------------------------
# Helpers for in-memory XLSX construction (T084, T085, T090)
# ---------------------------------------------------------------------------


def _build_simple_xlsx(headers: list, data_rows: list[list]) -> bytes:
    """Build an in-memory XLSX with the *simple* GHFDB layout.

    Layout:
      Rows 1-5: Arbitrary metadata (skipped by GHFDBSimpleImportFormat)
      Row 6:    Column headers
      Row 7+:   Data rows

    Args:
        headers: List of column header strings for row 6.
        data_rows: List of value lists, one per data row (row 7+).

    Returns:
        Raw bytes of the XLSX file.
    """
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data list"

    for i in range(1, 6):
        ws.cell(row=i, column=1, value=f"Metadata row {i}")

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=6, column=col_idx, value=header)

    for row_idx, row_values in enumerate(data_rows, start=7):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# T084 — GHFDBSimpleImportFormat unit tests
# ---------------------------------------------------------------------------


class TestGHFDBSimpleImportFormat:
    """T084 — Unit tests for GHFDBSimpleImportFormat."""

    def test_get_title_returns_simple_label(self):
        """get_title() returns 'GHFDB Simple Template'."""
        from project.ghfdb.resources import GHFDBSimpleImportFormat

        fmt = GHFDBSimpleImportFormat()
        assert fmt.get_title() == "GHFDB Simple Template"

    def test_official_format_get_title_returns_official_label(self):
        """GHFDBImportFormat.get_title() returns 'GHFDB Official Template'."""
        from project.ghfdb.resources import GHFDBImportFormat

        fmt = GHFDBImportFormat()
        assert fmt.get_title() == "GHFDB Official Template"

    @pytest.mark.xfail(strict=True, reason=(
        "Half-landed GHFDB canonical column work: the constants moved to the "
        "published spreadsheet casing, but ghfdb_colmeta.json, the resource "
        "field declarations and one manager annotation key did not follow. "
        "Needs debugging, and a decision on the published column vocabulary, "
        "before it can pass. See issue #122."
    ))
    def test_create_dataset_returns_correct_row_count(self):
        """create_dataset() with 2 data rows returns a Dataset with 2 rows."""
        from project.ghfdb.resources import GHFDBSimpleImportFormat
        from project.ghfdb.resources.formats import PARENT_COLUMNS

        xlsx_bytes = _build_simple_xlsx(
            headers=PARENT_COLUMNS,
            data_rows=[
                [
                    "GHFDB-P-001",
                    "70.0",
                    "",
                    "Site A",
                    "48.0",
                    "11.0",
                    "",
                    "Onshore (continental)",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "Germany",
                    "",
                    "Europe",
                    "",
                ],
                [
                    "GHFDB-P-002",
                    "80.0",
                    "",
                    "Site B",
                    "49.0",
                    "12.0",
                    "",
                    "Onshore (continental)",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "Germany",
                    "",
                    "Europe",
                    "",
                ],
            ],
        )

        fmt = GHFDBSimpleImportFormat()
        ds = fmt.create_dataset(xlsx_bytes)

        assert len(ds) == 2

    @pytest.mark.xfail(strict=True, reason=(
        "Half-landed GHFDB canonical column work: the constants moved to the "
        "published spreadsheet casing, but ghfdb_colmeta.json, the resource "
        "field declarations and one manager annotation key did not follow. "
        "Needs debugging, and a decision on the published column vocabulary, "
        "before it can pass. See issue #122."
    ))
    def test_create_dataset_uses_row6_as_headers(self):
        """create_dataset() uses row 6 cell values as Dataset column headers."""
        from project.ghfdb.resources import GHFDBSimpleImportFormat
        from project.ghfdb.resources.formats import PARENT_COLUMNS

        xlsx_bytes = _build_simple_xlsx(
            headers=PARENT_COLUMNS,
            data_rows=[["GHFDB-P-001"] + [""] * (len(PARENT_COLUMNS) - 1)],
        )

        fmt = GHFDBSimpleImportFormat()
        ds = fmt.create_dataset(xlsx_bytes)

        assert list(ds.headers) == PARENT_COLUMNS

    @pytest.mark.xfail(strict=True, reason=(
        "Half-landed GHFDB canonical column work: the constants moved to the "
        "published spreadsheet casing, but ghfdb_colmeta.json, the resource "
        "field declarations and one manager annotation key did not follow. "
        "Needs debugging, and a decision on the published column vocabulary, "
        "before it can pass. See issue #122."
    ))
    def test_create_dataset_excludes_metadata_rows(self):
        """create_dataset() does not include metadata rows 1-5 as data rows."""
        from project.ghfdb.resources import GHFDBSimpleImportFormat
        from project.ghfdb.resources.formats import PARENT_COLUMNS

        xlsx_bytes = _build_simple_xlsx(
            headers=PARENT_COLUMNS,
            data_rows=[["GHFDB-P-001"] + [""] * (len(PARENT_COLUMNS) - 1)],
        )

        fmt = GHFDBSimpleImportFormat()
        ds = fmt.create_dataset(xlsx_bytes)

        # Metadata rows 1-5 contain "Metadata row N" text; no data row should contain it
        flat_values = [str(v) for row in ds.dict for v in row.values() if v is not None]
        assert not any("Metadata row" in v for v in flat_values), (
            "Metadata row text found in dataset — metadata rows were not skipped"
        )

    @pytest.mark.xfail(strict=True, reason=(
        "Half-landed GHFDB canonical column work: the constants moved to the "
        "published spreadsheet casing, but ghfdb_colmeta.json, the resource "
        "field declarations and one manager annotation key did not follow. "
        "Needs debugging, and a decision on the published column vocabulary, "
        "before it can pass. See issue #122."
    ))
    def test_create_dataset_differs_from_official_format_on_offset_data(self):
        """Simple format returns data from row 7; official format skips rows 7-8 and gets data from row 9."""
        from io import BytesIO

        import openpyxl

        from project.ghfdb.resources import GHFDBImportFormat, GHFDBSimpleImportFormat
        from project.ghfdb.resources.formats import PARENT_COLUMNS

        # Build official-style XLSX with unit row (7), range row (8), then data row (9)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "data list"

        for i in range(1, 6):
            ws.cell(row=i, column=1, value=f"Metadata row {i}")
        for col_idx, header in enumerate(PARENT_COLUMNS, start=1):
            ws.cell(row=6, column=col_idx, value=header)
        ws.cell(row=7, column=1, value="UNIT_ROW")
        ws.cell(row=8, column=1, value="RANGE_ROW")
        ws.cell(row=9, column=1, value="GHFDB-P-001")

        buf = BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        official_ds = GHFDBImportFormat().create_dataset(xlsx_bytes)
        simple_ds = GHFDBSimpleImportFormat().create_dataset(xlsx_bytes)

        # Official format: 1 row (row 9 only)
        assert len(official_ds) == 1
        assert official_ds[0][0] == "GHFDB-P-001"

        # Simple format: 3 rows (rows 7, 8, 9 — including unit/range rows as data)
        assert len(simple_ds) == 3
        assert simple_ds[0][0] == "UNIT_ROW"


# ---------------------------------------------------------------------------
# T085 — Admin format-selection tests
# ---------------------------------------------------------------------------


class TestAdminGetImportFormats:
    """T085 — Admin classes expose both import format classes."""

    def test_child_admin_get_import_formats_returns_two_classes(self):
        """GHFDBChildAdmin.get_import_formats() returns [GHFDBImportFormat, GHFDBSimpleImportFormat]."""
        from django.contrib.admin import AdminSite

        from project.ghfdb.admin import GHFDBChildAdmin
        from project.ghfdb.models import GHFDBChild
        from project.ghfdb.resources import GHFDBImportFormat, GHFDBSimpleImportFormat

        site_admin = GHFDBChildAdmin(GHFDBChild, AdminSite())
        formats = site_admin.get_import_formats()

        assert len(formats) == 2
        assert formats[0] is GHFDBImportFormat
        assert formats[1] is GHFDBSimpleImportFormat

    def test_parent_admin_get_import_formats_returns_two_classes(self):
        """GHFDBParentAdmin.get_import_formats() returns [GHFDBImportFormat, GHFDBSimpleImportFormat]."""
        from django.contrib.admin import AdminSite

        from project.ghfdb.admin import GHFDBParentAdmin
        from project.ghfdb.models import GHFDBParent
        from project.ghfdb.resources import GHFDBImportFormat, GHFDBSimpleImportFormat

        site_admin = GHFDBParentAdmin(GHFDBParent, AdminSite())
        formats = site_admin.get_import_formats()

        assert len(formats) == 2
        assert formats[0] is GHFDBImportFormat
        assert formats[1] is GHFDBSimpleImportFormat

    def test_child_admin_format_titles(self):
        """Both format classes returned by GHFDBChildAdmin have the correct get_title() values."""
        from django.contrib.admin import AdminSite

        from project.ghfdb.admin import GHFDBChildAdmin
        from project.ghfdb.models import GHFDBChild

        site_admin = GHFDBChildAdmin(GHFDBChild, AdminSite())
        formats = site_admin.get_import_formats()

        assert formats[0]().get_title() == "GHFDB Official Template"
        assert formats[1]().get_title() == "GHFDB Simple Template"

    def test_parent_admin_format_titles(self):
        """Both format classes returned by GHFDBParentAdmin have the correct get_title() values."""
        from django.contrib.admin import AdminSite

        from project.ghfdb.admin import GHFDBParentAdmin
        from project.ghfdb.models import GHFDBParent

        site_admin = GHFDBParentAdmin(GHFDBParent, AdminSite())
        formats = site_admin.get_import_formats()

        assert formats[0]().get_title() == "GHFDB Official Template"
        assert formats[1]().get_title() == "GHFDB Simple Template"

    @pytest.mark.django_db
    def test_child_admin_import_page_shows_both_format_titles(self, admin_client):
        """GET admin import page for GHFDB child contains both format option texts."""
        from django.urls import reverse

        url = reverse("admin:ghfdb_ghfdbchild_import")
        response = admin_client.get(url)

        assert response.status_code == 200
        content = response.content.decode()
        assert "GHFDB Official Template" in content
        assert "GHFDB Simple Template" in content

    @pytest.mark.django_db
    def test_parent_admin_import_page_shows_both_format_titles(self, admin_client):
        """GET admin import page for GHFDBParent contains both format option texts."""
        from django.urls import reverse

        url = reverse("admin:ghfdb_ghfdbparent_import")
        response = admin_client.get(url)

        assert response.status_code == 200
        content = response.content.decode()
        assert "GHFDB Official Template" in content
        assert "GHFDB Simple Template" in content


@pytest.mark.django_db
class TestGHFDBParentPrivateDatasetRegression:
    """The parent import must find its target dataset even when that dataset is private."""

    def test_import_attaches_records_to_a_private_dataset(self, dataset):
        """A row imports into the only dataset available, whether or not it is public."""
        from fairdm.utils.choices import Visibility
        from heat_flow.models import HeatFlowSite, ParentHeatFlow

        from project.ghfdb.resources import GHFDBParentImportResource

        assert dataset.visibility == Visibility.PRIVATE

        resource = GHFDBParentImportResource()
        result = resource.import_data(make_dataset(PARENT_ROW), dry_run=False, raise_errors=False)

        assert not result.has_errors(), [
            (line, [str(e.error) for e in errors]) for line, errors in result.row_errors()
        ]
        assert ParentHeatFlow.objects.get(ghfdb_id=1).dataset == dataset
        assert HeatFlowSite.objects.get(name="Test Site Alpha").dataset == dataset


class TestGHFDBParentImportRecordsIdentifierOnMatchedSite:
    """A site matched by coordinates keeps the row's identifier for later lookups.

    Coordinates decide identity, so a row matching an existing site by coordinates
    binds to it.  If that site carries no identifier of its own, the row's is
    recorded: a later row carrying the same identifier and no coordinates would
    otherwise fail its lookup and create a duplicate.
    """

    @pytest.mark.django_db
    def test_identifier_is_recorded_on_a_site_matched_by_coordinates(self, dataset):
        """The site must pre-exist without an identifier, or the import takes the
        other branch and this passes whether or not the behaviour is there."""
        from decimal import Decimal

        from fairdm.contrib.location.models import Point
        from heat_flow.models import HeatFlowSite

        from project.ghfdb.resources import GHFDBParentImportResource

        point, _ = Point.objects.get_or_create(
            x=Decimal(PARENT_ROW["long_EW"]), y=Decimal(PARENT_ROW["lat_NS"])
        )
        existing = HeatFlowSite.objects.create(
            dataset=dataset, name="Entered by hand", location=point
        )
        assert not existing.local_id

        result = GHFDBParentImportResource().import_data(
            make_dataset(PARENT_ROW), dry_run=False, raise_errors=False
        )
        assert not result.has_errors(), result.invalid_rows

        # One site, not two: the row resolved to the one already there.
        assert HeatFlowSite.objects.filter(location=point).count() == 1
        existing.refresh_from_db()
        assert existing.local_id == PARENT_ROW["ID_parent"], (
            "The row's identifier was not recorded on the site it resolved to, so a "
            "later row carrying it without coordinates would create a duplicate."
        )
