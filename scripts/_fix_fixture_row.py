"""One-shot script: insert an 'Allowed range of values' row at row 8 of the
roundtrip test fixture so data starts at row 9, matching the real GHFDB
template layout."""

import openpyxl

FIXTURE = "tests/test_ghfdb/fixtures/sample_ghfdb.xlsx"

wb = openpyxl.load_workbook(FIXTURE)
ws = wb["data list"]

print("Before insert:")
for i in range(6, 12):
    print(f"  Row {i}: {[ws.cell(row=i, column=j).value for j in range(1, 5)]}")

# Insert a new blank row 8 — shifts current rows 8+ down by 1
ws.insert_rows(8, amount=1)

# Populate with 'Allowed range of values' placeholders based on row-6 headers
for c in range(1, ws.max_column + 1):
    hdr = ws.cell(row=6, column=c).value
    ws.cell(row=8, column=c).value = ("allowed range: " + str(hdr)) if hdr else None

print("\nAfter insert:")
for i in range(6, 13):
    print(f"  Row {i}: {[ws.cell(row=i, column=j).value for j in range(1, 5)]}")

wb.save(FIXTURE)
print(f"\nSaved {FIXTURE}")
