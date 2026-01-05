"""
Script to create test fixtures for the Global Heat Flow Database testing infrastructure.
This script generates Excel files conforming to the GHFDB template structure.
"""

from datetime import datetime
from pathlib import Path

import openpyxl

# Define the base path for fixtures
FIXTURES_DIR = Path(__file__).parent.parent / "tests" / "fixtures"
FIXTURES_DIR.mkdir(exist_ok=True)


def create_ghfdb_template():
    """Create a new Excel workbook with the GHFDB template structure."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data list"

    # Row 1-5: Template header section (simplified for tests)
    ws["A1"] = "Global Heat Flow Database Template"
    ws["A2"] = "Test Fixture Data"
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A4"] = ""
    ws["A5"] = ""

    # Row 6: Column headers (as per GHFDB specification)
    headers = [
        "Short Name",
        "q",
        "q_uncertainty",
        "name",
        "lat_NS",
        "long_EW",
        "elevation",
        "environment",
        "p_comment",
        "corr_HP_flag",
        "total_depth_MD",
        "total_depth_TVD",
        "explo_method",
        "explo_purpose",
        "qc",
        "qc_uncertainty",
        "q_method",
        "q_top",
        "q_bottom",
        "probe_penetration",
        "publication_reference",
        "data_reference",
        "relevant_child",
        "c_comment",
        "corr_IS_flag",
        "corr_T_flag",
        "corr_S_flag",
        "corr_E_flag",
        "corr_TOPO_flag",
        "corr_PAL_flag",
        "corr_SUR_flag",
        "corr_CONV_flag",
        "corr_HR_flag",
        "expedition",
        "probe_type",
        "probe_length",
        "probe_tilt",
        "water_temperature",
        "geo_lithology",
        "geo_stratigraphy",
        "T_grad_mean",
        "T_grad_uncertainty",
        "T_grad_mean_cor",
        "T_grad_uncertainty_cor",
        "T_method_top",
        "T_method_bottom",
        "T_shutin_top",
        "T_shutin_bottom",
        "T_corr_top",
        "T_corr_bottom",
        "T_number",
        "q_date",
        "tc_mean",
        "tc_uncertainty",
        "tc_source",
        "tc_location",
        "tc_method",
        "tc_saturation",
        "tc_pT_conditions",
        "tc_pT_function",
        "tc_number",
        "tc_strategy",
        "Ref_ISGN",
        "Reviewer_name",
        "Reviewer_comment",
        "Review_date",
        "Country",
        "Region",
        "Continent",
        "Domain",
        "ID",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=6, column=col_idx, value=header)

    # Row 7-8: Empty (template formatting rows)

    return wb


def create_minimal_fixture():
    """
    Create minimal_ghfdb_import.xlsx: 5 heat flow sites with complete mandatory fields.

    Sites represent diverse geographic locations and environments:
    1. Offshore continental drilling (North Atlantic)
    2. Onshore borehole (US continental)
    3. Marine probe (Pacific Ocean)
    4. Continental drilling (Central Europe)
    5. Geothermal exploration (Iceland)
    """
    wb = create_ghfdb_template()
    ws = wb.active

    # Data rows starting at row 9
    sites = [
        # Site 1: Offshore continental drilling (based on importer_success example)
        {
            "Short Name": 1,
            "q": 45.0,
            "q_uncertainty": 5.0,
            "name": "Offshore Atlantic Site",
            "lat_NS": 43.63333,
            "long_EW": -5.61667,
            "elevation": 67.0,
            "environment": "[Offshore (continental)]",
            "p_comment": None,
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 105.0,
            "total_depth_TVD": 100.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Hydrocarbon]",
            "qc": 63.0,
            "qc_uncertainty": 3.6,
            "q_method": "[Interval method]",
            "q_top": 10.0,
            "q_bottom": 45.0,
            "probe_penetration": 6.5,
            "relevant_child": "[Yes]",
            "c_comment": None,
            "corr_IS_flag": "[not considered]",
            "T_grad_mean": 32.0,
            "T_grad_uncertainty": 2.5,
            "T_method_top": "[BHT]",
            "T_method_bottom": "[BHT]",
            "tc_mean": 2.1,
            "tc_uncertainty": 0.15,
            "tc_method": "[Lab - divided bar]",
            "tc_source": "[Publication]",
            "Country": "Spain",
            "Region": "Bay of Biscay",
            "Continent": "Europe",
        },
        # Site 2: Onshore borehole (US continental)
        {
            "Short Name": 2,
            "q": 58.0,
            "q_uncertainty": 4.2,
            "name": "Texas Onshore Well",
            "lat_NS": 32.75,
            "long_EW": -97.33,
            "elevation": 200.0,
            "environment": "[Continental (plateau/cratonic)]",
            "corr_HP_flag": "[No]",
            "total_depth_MD": 2500.0,
            "total_depth_TVD": 2480.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Scientific]",
            "qc": 61.0,
            "qc_uncertainty": 4.5,
            "q_method": "[Bullard method]",
            "q_top": 500.0,
            "q_bottom": 2400.0,
            "relevant_child": "[Yes]",
            "corr_T_flag": "[paleoclimate]",
            "T_grad_mean": 28.5,
            "T_grad_uncertainty": 2.0,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "tc_mean": 2.15,
            "tc_uncertainty": 0.12,
            "tc_method": "[Lab - TCS]",
            "tc_source": "[Database]",
            "Country": "United States",
            "Region": "Texas",
            "Continent": "North America",
        },
        # Site 3: Marine probe (Pacific Ocean)
        {
            "Short Name": 3,
            "q": 72.0,
            "q_uncertainty": 8.0,
            "name": "Pacific Abyssal Probe",
            "lat_NS": -12.45,
            "long_EW": -110.25,
            "elevation": -3200.0,
            "environment": "[Offshore (oceanic)]",
            "corr_HP_flag": "[No]",
            "total_depth_MD": 5.0,
            "total_depth_TVD": 5.0,
            "explo_method": "[Probe]",
            "explo_purpose": "[Scientific]",
            "qc": 72.0,
            "qc_uncertainty": 8.0,
            "q_method": "[Probe]",
            "q_top": 0.5,
            "q_bottom": 4.5,
            "probe_penetration": 4.5,
            "probe_type": "[Lister-type]",
            "probe_length": 5.0,
            "water_temperature": 2.5,
            "relevant_child": "[Yes]",
            "T_grad_mean": 85.0,
            "T_grad_uncertainty": 10.0,
            "T_method_top": "[Thermistor]",
            "T_method_bottom": "[Thermistor]",
            "tc_mean": 0.85,
            "tc_uncertainty": 0.08,
            "tc_method": "[Lab - needle probe]",
            "tc_source": "[Measurement]",
            "tc_location": "[In situ]",
            "Country": None,
            "Region": "East Pacific Rise",
            "Continent": "Pacific Ocean",
        },
        # Site 4: Continental drilling (Central Europe)
        {
            "Short Name": 4,
            "q": 52.0,
            "q_uncertainty": 3.5,
            "name": "German Continental Well",
            "lat_NS": 51.23,
            "long_EW": 10.45,
            "elevation": 350.0,
            "environment": "[Continental (orogenic)]",
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 4200.0,
            "total_depth_TVD": 4150.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Scientific]",
            "qc": 55.0,
            "qc_uncertainty": 3.8,
            "q_method": "[Interval method]",
            "q_top": 800.0,
            "q_bottom": 3900.0,
            "relevant_child": "[Yes]",
            "corr_TOPO_flag": "[topography]",
            "geo_lithology": "[Granite]",
            "geo_stratigraphy": "[Paleozoic]",
            "T_grad_mean": 26.0,
            "T_grad_uncertainty": 1.8,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "tc_mean": 2.8,
            "tc_uncertainty": 0.2,
            "tc_method": "[Lab - TCS]",
            "tc_source": "[Measurement]",
            "tc_location": "[Lab]",
            "Country": "Germany",
            "Region": "Thuringian Basin",
            "Continent": "Europe",
        },
        # Site 5: Geothermal exploration (Iceland)
        {
            "Short Name": 5,
            "q": 125.0,
            "q_uncertainty": 12.0,
            "name": "Iceland Geothermal",
            "lat_NS": 64.15,
            "long_EW": -21.95,
            "elevation": 180.0,
            "environment": "[Continental (volcanic)]",
            "corr_HP_flag": "[No]",
            "total_depth_MD": 1500.0,
            "total_depth_TVD": 1480.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Geothermal]",
            "qc": 128.0,
            "qc_uncertainty": 13.0,
            "q_method": "[Bullard method]",
            "q_top": 200.0,
            "q_bottom": 1400.0,
            "relevant_child": "[Yes]",
            "corr_CONV_flag": "[groundwater]",
            "geo_lithology": "[Basalt]",
            "geo_stratigraphy": "[Quaternary]",
            "T_grad_mean": 90.0,
            "T_grad_uncertainty": 8.0,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "tc_mean": 1.45,
            "tc_uncertainty": 0.1,
            "tc_method": "[Lab - divided bar]",
            "tc_source": "[Measurement]",
            "Country": "Iceland",
            "Region": "Reykjanes Peninsula",
            "Continent": "Europe",
        },
    ]

    # Write data rows starting at row 9
    for row_idx, site_data in enumerate(sites, start=9):
        # Get headers from row 6
        headers = [cell.value for cell in ws[6] if cell.value]

        for col_idx, header in enumerate(headers, start=1):
            value = site_data.get(header)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook
    output_path = FIXTURES_DIR / "minimal_ghfdb_import.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_invalid_fixture():
    """
    Create invalid_ghfdb_import.xlsx: 5 validation error cases.

    Each row demonstrates a different validation failure:
    1. Missing site name (mandatory field)
    2. Latitude > 90 (out of bounds)
    3. Negative heat flow value (invalid)
    4. Empty mandatory environment field
    5. Depth values inconsistent (top > bottom)
    """
    wb = create_ghfdb_template()
    ws = wb.active

    invalid_sites = [
        # Case 1: Missing name (mandatory field)
        {
            "Short Name": 101,
            "q": 45.0,
            "q_uncertainty": 5.0,
            "name": None,  # INVALID: mandatory field missing
            "lat_NS": 40.0,
            "long_EW": -75.0,
            "elevation": 100.0,
            "environment": "[Continental (plateau/cratonic)]",
            "explo_method": "[Drilling]",
            "qc": 48.0,
            "qc_uncertainty": 4.0,
            "q_method": "[Interval method]",
            "q_top": 100.0,
            "q_bottom": 500.0,
            "relevant_child": "[Yes]",
            "T_grad_mean": 25.0,
            "T_grad_uncertainty": 2.0,
            "tc_mean": 2.0,
            "tc_uncertainty": 0.15,
        },
        # Case 2: Latitude out of bounds
        {
            "Short Name": 102,
            "q": 50.0,
            "q_uncertainty": 4.0,
            "name": "Invalid Latitude Site",
            "lat_NS": 95.0,  # INVALID: latitude > 90
            "long_EW": 10.0,
            "elevation": 200.0,
            "environment": "[Continental (plateau/cratonic)]",
            "explo_method": "[Drilling]",
            "qc": 52.0,
            "qc_uncertainty": 4.5,
            "q_method": "[Interval method]",
            "q_top": 100.0,
            "q_bottom": 500.0,
            "relevant_child": "[Yes]",
            "T_grad_mean": 26.0,
            "T_grad_uncertainty": 2.5,
            "tc_mean": 2.1,
            "tc_uncertainty": 0.12,
        },
        # Case 3: Negative heat flow
        {
            "Short Name": 103,
            "q": -25.0,  # INVALID: negative heat flow
            "q_uncertainty": 3.0,
            "name": "Negative Heat Flow Site",
            "lat_NS": 35.0,
            "long_EW": -120.0,
            "elevation": 50.0,
            "environment": "[Continental (plateau/cratonic)]",
            "explo_method": "[Drilling]",
            "qc": -28.0,  # INVALID: negative
            "qc_uncertainty": 3.5,
            "q_method": "[Interval method]",
            "q_top": 200.0,
            "q_bottom": 800.0,
            "relevant_child": "[Yes]",
            "T_grad_mean": 20.0,
            "T_grad_uncertainty": 2.0,
            "tc_mean": 2.2,
            "tc_uncertainty": 0.18,
        },
        # Case 4: Missing environment (mandatory field)
        {
            "Short Name": 104,
            "q": 55.0,
            "q_uncertainty": 5.5,
            "name": "No Environment Site",
            "lat_NS": 48.0,
            "long_EW": 15.0,
            "elevation": 300.0,
            "environment": None,  # INVALID: mandatory field missing
            "explo_method": "[Drilling]",
            "qc": 58.0,
            "qc_uncertainty": 6.0,
            "q_method": "[Interval method]",
            "q_top": 150.0,
            "q_bottom": 600.0,
            "relevant_child": "[Yes]",
            "T_grad_mean": 27.0,
            "T_grad_uncertainty": 2.3,
            "tc_mean": 2.15,
            "tc_uncertainty": 0.14,
        },
        # Case 5: Inconsistent depth values (top > bottom)
        {
            "Short Name": 105,
            "q": 62.0,
            "q_uncertainty": 6.0,
            "name": "Invalid Depth Site",
            "lat_NS": 52.0,
            "long_EW": -2.0,
            "elevation": 150.0,
            "environment": "[Continental (plateau/cratonic)]",
            "explo_method": "[Drilling]",
            "qc": 65.0,
            "qc_uncertainty": 6.5,
            "q_method": "[Interval method]",
            "q_top": 1000.0,  # INVALID: top > bottom
            "q_bottom": 500.0,
            "relevant_child": "[Yes]",
            "T_grad_mean": 29.0,
            "T_grad_uncertainty": 2.5,
            "tc_mean": 2.25,
            "tc_uncertainty": 0.16,
        },
    ]

    # Write data rows starting at row 9
    for row_idx, site_data in enumerate(invalid_sites, start=9):
        headers = [cell.value for cell in ws[6] if cell.value]

        for col_idx, header in enumerate(headers, start=1):
            value = site_data.get(header)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook
    output_path = FIXTURES_DIR / "invalid_ghfdb_import.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_round_trip_fixture():
    """
    Create round_trip_reference.xlsx: Comprehensive fixture with 10 sites.

    Tests complete export→import roundtrip integrity with:
    - All environment types
    - All exploration methods and purposes
    - All measurement types (probe, borehole, lab)
    - All correction flags
    - Optional fields (probe details, water temp, etc.)
    - Complete controlled vocabularies
    """
    wb = create_ghfdb_template()
    ws = wb.active

    # 10 comprehensive sites covering all field variations
    sites = [
        # Site 1: Offshore continental with ALL correction flags
        {
            "Short Name": 1001,
            "q": 48.5,
            "q_uncertainty": 4.2,
            "name": "Comprehensive Site 1 - All Corrections",
            "lat_NS": 45.123,
            "long_EW": -3.456,
            "elevation": -125.5,
            "environment": "[Offshore (continental)]",
            "p_comment": "Test site with all correction flags enabled",
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 3500.0,
            "total_depth_TVD": 3450.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Hydrocarbon]",
            "qc": 52.3,
            "qc_uncertainty": 4.8,
            "q_method": "[Bullard method]",
            "q_top": 500.0,
            "q_bottom": 3200.0,
            "relevant_child": "[Yes]",
            "c_comment": "All corrections applied",
            "corr_IS_flag": "[steady state]",
            "corr_T_flag": "[paleoclimate]",
            "corr_S_flag": "[sedimentation]",
            "corr_E_flag": "[erosion]",
            "corr_TOPO_flag": "[topography]",
            "corr_PAL_flag": "[paleoclimate]",
            "corr_SUR_flag": "[surface water]",
            "corr_CONV_flag": "[groundwater]",
            "corr_HR_flag": "[heat refraction]",
            "geo_lithology": "[Sandstone]",
            "geo_stratigraphy": "[Cenozoic]",
            "T_grad_mean": 28.5,
            "T_grad_uncertainty": 2.1,
            "T_method_top": "[BHT]",
            "T_method_bottom": "[BHT]",
            "T_shutin_top": 24.0,
            "T_shutin_bottom": 48.0,
            "T_number": 15,
            "q_date": "2024-05-15",
            "tc_mean": 2.35,
            "tc_uncertainty": 0.18,
            "tc_source": "[Measurement]",
            "tc_location": "[Lab]",
            "tc_method": "[Lab - TCS]",
            "tc_saturation": "[Saturated]",
            "tc_pT_conditions": "[In situ]",
            "tc_number": 25,
            "tc_strategy": "[Averaging]",
            "Country": "France",
            "Region": "Bay of Biscay",
            "Continent": "Europe",
            "Domain": "Aquitaine Basin",
        },
        # Site 2: Oceanic with probe measurements
        {
            "Short Name": 1002,
            "q": 95.0,
            "q_uncertainty": 12.0,
            "name": "Deep Ocean Probe Site",
            "lat_NS": -25.678,
            "long_EW": 145.234,
            "elevation": -4500.0,
            "environment": "[Offshore (oceanic)]",
            "corr_HP_flag": "[No]",
            "total_depth_MD": 8.5,
            "total_depth_TVD": 8.5,
            "explo_method": "[Probe]",
            "explo_purpose": "[Scientific]",
            "qc": 95.0,
            "qc_uncertainty": 12.0,
            "q_method": "[Probe]",
            "q_top": 1.0,
            "q_bottom": 7.5,
            "probe_penetration": 7.5,
            "relevant_child": "[Yes]",
            "probe_type": "[Ewing-type]",
            "probe_length": 8.5,
            "probe_tilt": 2.3,
            "water_temperature": 1.8,
            "expedition": "RV Atlantis Expedition 2024",
            "T_grad_mean": 110.0,
            "T_grad_uncertainty": 15.0,
            "T_method_top": "[Thermistor]",
            "T_method_bottom": "[Thermistor]",
            "tc_mean": 0.88,
            "tc_uncertainty": 0.09,
            "tc_source": "[Measurement]",
            "tc_location": "[In situ]",
            "tc_method": "[Lab - needle probe]",
            "Country": None,
            "Region": "Tasman Sea",
            "Continent": "Pacific Ocean",
        },
        # Site 3: Continental plateau - cratonic
        {
            "Short Name": 1003,
            "q": 42.0,
            "q_uncertainty": 3.5,
            "name": "Stable Craton Site",
            "lat_NS": 62.45,
            "long_EW": 95.12,
            "elevation": 125.0,
            "environment": "[Continental (plateau/cratonic)]",
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 5200.0,
            "total_depth_TVD": 5150.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Scientific]",
            "qc": 44.5,
            "qc_uncertainty": 3.8,
            "q_method": "[Interval method]",
            "q_top": 1000.0,
            "q_bottom": 5000.0,
            "relevant_child": "[Yes]",
            "corr_T_flag": "[paleoclimate]",
            "geo_lithology": "[Gneiss]",
            "geo_stratigraphy": "[Precambrian]",
            "T_grad_mean": 18.5,
            "T_grad_uncertainty": 1.2,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "T_number": 20,
            "tc_mean": 2.95,
            "tc_uncertainty": 0.22,
            "tc_source": "[Database]",
            "tc_location": "[Lab]",
            "tc_method": "[Lab - divided bar]",
            "tc_number": 30,
            "Country": "Russia",
            "Region": "Siberian Platform",
            "Continent": "Asia",
            "Domain": "Siberian Craton",
        },
        # Site 4: Orogenic belt
        {
            "Short Name": 1004,
            "q": 68.0,
            "q_uncertainty": 6.5,
            "name": "Active Orogen Site",
            "lat_NS": 38.92,
            "long_EW": 22.15,
            "elevation": 850.0,
            "environment": "[Continental (orogenic)]",
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 2800.0,
            "total_depth_TVD": 2750.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Geothermal]",
            "qc": 72.0,
            "qc_uncertainty": 7.0,
            "q_method": "[Interval method]",
            "q_top": 300.0,
            "q_bottom": 2600.0,
            "relevant_child": "[Yes]",
            "corr_TOPO_flag": "[topography]",
            "corr_CONV_flag": "[groundwater]",
            "geo_lithology": "[Schist]",
            "geo_stratigraphy": "[Mesozoic]",
            "T_grad_mean": 35.0,
            "T_grad_uncertainty": 3.2,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "T_corr_top": 2.5,
            "T_corr_bottom": 3.0,
            "tc_mean": 2.1,
            "tc_uncertainty": 0.15,
            "tc_source": "[Measurement]",
            "tc_location": "[Lab]",
            "tc_method": "[Lab - TCS]",
            "Country": "Greece",
            "Region": "Hellenic Arc",
            "Continent": "Europe",
        },
        # Site 5: Volcanic setting
        {
            "Short Name": 1005,
            "q": 150.0,
            "q_uncertainty": 18.0,
            "name": "Volcanic Arc Site",
            "lat_NS": -15.35,
            "long_EW": -173.25,
            "elevation": 320.0,
            "environment": "[Continental (volcanic)]",
            "corr_HP_flag": "[No]",
            "total_depth_MD": 1800.0,
            "total_depth_TVD": 1780.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Geothermal]",
            "qc": 155.0,
            "qc_uncertainty": 19.0,
            "q_method": "[Bullard method]",
            "q_top": 400.0,
            "q_bottom": 1700.0,
            "relevant_child": "[Yes]",
            "corr_CONV_flag": "[groundwater]",
            "geo_lithology": "[Andesite]",
            "geo_stratigraphy": "[Quaternary]",
            "T_grad_mean": 120.0,
            "T_grad_uncertainty": 12.0,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "tc_mean": 1.28,
            "tc_uncertainty": 0.11,
            "tc_source": "[Measurement]",
            "tc_location": "[Lab]",
            "tc_method": "[Lab - TCS]",
            "Country": "Samoa",
            "Region": "Samoa Volcanic Chain",
            "Continent": "Pacific Ocean",
        },
        # Site 6: Rift valley
        {
            "Short Name": 1006,
            "q": 105.0,
            "q_uncertainty": 10.0,
            "name": "Continental Rift Site",
            "lat_NS": -2.45,
            "long_EW": 36.82,
            "elevation": 1200.0,
            "environment": "[Continental (orogenic)]",
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 3200.0,
            "total_depth_TVD": 3180.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Scientific]",
            "qc": 108.0,
            "qc_uncertainty": 10.5,
            "q_method": "[Interval method]",
            "q_top": 600.0,
            "q_bottom": 3000.0,
            "relevant_child": "[Yes]",
            "corr_TOPO_flag": "[topography]",
            "geo_lithology": "[Basalt]",
            "geo_stratigraphy": "[Cenozoic]",
            "T_grad_mean": 48.0,
            "T_grad_uncertainty": 4.5,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "tc_mean": 2.25,
            "tc_uncertainty": 0.18,
            "tc_source": "[Measurement]",
            "tc_location": "[Lab]",
            "tc_method": "[Lab - divided bar]",
            "Country": "Kenya",
            "Region": "East African Rift",
            "Continent": "Africa",
        },
        # Site 7: Marine sedimentary basin
        {
            "Short Name": 1007,
            "q": 58.0,
            "q_uncertainty": 5.2,
            "name": "Marine Sedimentary Basin",
            "lat_NS": 28.65,
            "long_EW": -92.35,
            "elevation": -1850.0,
            "environment": "[Offshore (continental)]",
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 4200.0,
            "total_depth_TVD": 4150.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Hydrocarbon]",
            "qc": 61.0,
            "qc_uncertainty": 5.6,
            "q_method": "[Bullard method]",
            "q_top": 800.0,
            "q_bottom": 4000.0,
            "relevant_child": "[Yes]",
            "corr_S_flag": "[sedimentation]",
            "corr_HR_flag": "[heat refraction]",
            "geo_lithology": "[Shale]",
            "geo_stratigraphy": "[Cenozoic]",
            "T_grad_mean": 29.5,
            "T_grad_uncertainty": 2.4,
            "T_method_top": "[BHT]",
            "T_method_bottom": "[BHT]",
            "T_shutin_top": 12.0,
            "T_shutin_bottom": 18.0,
            "tc_mean": 2.08,
            "tc_uncertainty": 0.16,
            "tc_source": "[Publication]",
            "tc_location": "[Lab]",
            "tc_method": "[Lab - TCS]",
            "Country": "United States",
            "Region": "Gulf of Mexico",
            "Continent": "North America",
        },
        # Site 8: Onshore hydrocarbon well
        {
            "Short Name": 1008,
            "q": 54.0,
            "q_uncertainty": 4.8,
            "name": "Onshore Hydrocarbon Well",
            "lat_NS": 31.25,
            "long_EW": 48.92,
            "elevation": 45.0,
            "environment": "[Continental (plateau/cratonic)]",
            "corr_HP_flag": "[No]",
            "total_depth_MD": 3800.0,
            "total_depth_TVD": 3750.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Hydrocarbon]",
            "qc": 56.5,
            "qc_uncertainty": 5.1,
            "q_method": "[Interval method]",
            "q_top": 1200.0,
            "q_bottom": 3600.0,
            "relevant_child": "[Yes]",
            "corr_IS_flag": "[steady state]",
            "geo_lithology": "[Limestone]",
            "geo_stratigraphy": "[Mesozoic]",
            "T_grad_mean": 26.0,
            "T_grad_uncertainty": 2.0,
            "T_method_top": "[BHT]",
            "T_method_bottom": "[BHT]",
            "T_shutin_top": 24.0,
            "T_shutin_bottom": 36.0,
            "tc_mean": 2.18,
            "tc_uncertainty": 0.17,
            "tc_source": "[Database]",
            "tc_method": "[Lab - divided bar]",
            "Country": "Iraq",
            "Region": "Mesopotamian Basin",
            "Continent": "Asia",
        },
        # Site 9: Shallow onshore site
        {
            "Short Name": 1009,
            "q": 38.0,
            "q_uncertainty": 3.2,
            "name": "Shallow Scientific Borehole",
            "lat_NS": 52.12,
            "long_EW": -106.63,
            "elevation": 525.0,
            "environment": "[Continental (plateau/cratonic)]",
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 250.0,
            "total_depth_TVD": 248.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Scientific]",
            "qc": 40.0,
            "qc_uncertainty": 3.5,
            "q_method": "[Interval method]",
            "q_top": 50.0,
            "q_bottom": 230.0,
            "relevant_child": "[Yes]",
            "corr_T_flag": "[paleoclimate]",
            "geo_lithology": "[Sandstone]",
            "geo_stratigraphy": "[Cenozoic]",
            "T_grad_mean": 22.0,
            "T_grad_uncertainty": 1.8,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "tc_mean": 1.82,
            "tc_uncertainty": 0.14,
            "tc_source": "[Measurement]",
            "tc_location": "[Lab]",
            "tc_method": "[Lab - TCS]",
            "Country": "Canada",
            "Region": "Saskatchewan",
            "Continent": "North America",
        },
        # Site 10: Mountain site with topographic correction
        {
            "Short Name": 1010,
            "q": 62.0,
            "q_uncertainty": 5.8,
            "name": "Alpine Mountain Site",
            "lat_NS": 46.85,
            "long_EW": 8.63,
            "elevation": 2450.0,
            "environment": "[Continental (orogenic)]",
            "corr_HP_flag": "[Yes]",
            "total_depth_MD": 1500.0,
            "total_depth_TVD": 1480.0,
            "explo_method": "[Drilling]",
            "explo_purpose": "[Scientific]",
            "qc": 65.0,
            "qc_uncertainty": 6.2,
            "q_method": "[Bullard method]",
            "q_top": 200.0,
            "q_bottom": 1400.0,
            "relevant_child": "[Yes]",
            "corr_TOPO_flag": "[topography]",
            "geo_lithology": "[Granite]",
            "geo_stratigraphy": "[Paleozoic]",
            "T_grad_mean": 30.0,
            "T_grad_uncertainty": 2.8,
            "T_method_top": "[Logging]",
            "T_method_bottom": "[Logging]",
            "T_corr_top": 3.5,
            "T_corr_bottom": 4.0,
            "tc_mean": 2.68,
            "tc_uncertainty": 0.20,
            "tc_source": "[Measurement]",
            "tc_location": "[Lab]",
            "tc_method": "[Lab - TCS]",
            "Country": "Switzerland",
            "Region": "Swiss Alps",
            "Continent": "Europe",
            "Domain": "Alpine Orogen",
        },
    ]

    # Write data rows starting at row 9
    for row_idx, site_data in enumerate(sites, start=9):
        headers = [cell.value for cell in ws[6] if cell.value]

        for col_idx, header in enumerate(headers, start=1):
            value = site_data.get(header)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook
    output_path = FIXTURES_DIR / "round_trip_reference.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


if __name__ == "__main__":
    print("Creating GHFDB test fixtures...")
    print("=" * 60)

    # Create fixtures
    create_minimal_fixture()
    create_invalid_fixture()
    create_round_trip_fixture()

    print("=" * 60)
    print("Fixture creation complete!")
