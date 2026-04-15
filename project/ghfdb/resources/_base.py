"""
Shared constants and base format class for GHFDB import/export resources.

Defines:
- GHFDBImportFormat: Custom XLSX reader for the GHFDB spreadsheet template
- GHFDB_COLUMN_ORDER: Canonical 62-column order matching ghfdb_colmeta.json
- PARENT_COLUMNS: 18 parent-level spreadsheet column names
- CORRECTION_COL_MAP: Mapping of correction flag columns to CorrectionTypeChoices

References:
    - Fuchs et al. (2021). A new database structure for the IHFC Global Heat
      Flow Database. Earth System Science Data.
    - Fuchs et al. (2023). The Global Heat Flow Database: Update 2023.
"""

from io import BytesIO

import tablib
from import_export.formats.base_formats import XLSX

# ---------------------------------------------------------------------------
# GHFDB_COLUMN_ORDER
# 62 canonical GHFDB column names, matching keys in ghfdb_colmeta.json.
# Used as Meta.export_order in GHFDBExportResource to enforce the canonical
# GHFDB spreadsheet column sequence.
# ---------------------------------------------------------------------------
GHFDB_COLUMN_ORDER: tuple[str, ...] = (
    "q",
    "q_uncertainty",
    "name",
    "lat_ns",
    "long_ew",
    "elevation",
    "environment",
    "p_comment",
    "corr_hp_flag",
    "total_depth_md",
    "total_depth_tvd",
    "explo_method",
    "explo_purpose",
    "qc",
    "qc_uncertainty",
    "q_method",
    "q_top",
    "q_bottom",
    "probe_penetration",
    "publication_reference",
    "data_reference",
    "relevant_child",
    "c_comment",
    "corr_is_flag",
    "corr_t_flag",
    "corr_s_flag",
    "corr_e_flag",
    "corr_topo_flag",
    "corr_pal_flag",
    "corr_sur_flag",
    "corr_conv_flag",
    "corr_hr_flag",
    "expedition",
    "probe_type",
    "probe_length",
    "probe_tilt",
    "water_temperature",
    "geo_lithology",
    "geo_stratigraphy",
    "t_grad_mean",
    "t_grad_uncertainty",
    "t_grad_mean_cor",
    "t_grad_uncertainty_cor",
    "t_method_top",
    "t_method_bottom",
    "t_shutin_top",
    "t_shutin_bottom",
    "t_corr_top",
    "t_corr_bottom",
    "t_number",
    "q_date",
    "tc_mean",
    "tc_uncertainty",
    "tc_source",
    "tc_location",
    "tc_method",
    "tc_saturation",
    "tc_pT_conditions",
    "tc_pT_function",
    "tc_number",
    "tc_strategy",
    "igsn",
)

# ---------------------------------------------------------------------------
# PARENT_COLUMNS
# The 18 spreadsheet column names that belong to the parent level.
# Used by GHFDBParentImportResource.before_import() to deduplicate rows and
# extract unique parent records from the flat GHFDB XLSX.
# Note: Uses actual spreadsheet header names (case-sensitive as in row 6 of
# the official GHFDB XLSX template).
# ---------------------------------------------------------------------------
PARENT_COLUMNS: list[str] = [
    "ID_parent",
    "q",
    "q_uncertainty",
    "name",
    "lat_NS",
    "long_EW",
    "elevation",
    "environment",
    "p_comment",
    "corr_HP_flag",
    "total_depth_MD",
    "total_depth_TVD",
    "explo_method",
    "explo_purpose",
    "Country",
    "Region",
    "Continent",
    "Domain",
]

# ---------------------------------------------------------------------------
# CORRECTION_COL_MAP
# Maps the 9 GHFDB correction-flag spreadsheet column headers to the
# corresponding HeatFlowCorrection.CorrectionTypeChoices value.
# Keys use actual spreadsheet header names; values are TextChoices values.
# ---------------------------------------------------------------------------
CORRECTION_COL_MAP: dict[str, str] = {
    "corr_IS_flag": "IS",
    "corr_T_flag": "T",
    "corr_S_flag": "S",
    "corr_E_flag": "E",
    "corr_TOPO_flag": "TOPO",
    "corr_PAL_flag": "PAL",
    "corr_SUR_flag": "SUR",
    "corr_CONV_flag": "CONV",
    "corr_HR_flag": "HR",
}


class GHFDBImportFormat(XLSX):
    """Custom XLSX format for the GHFDB spreadsheet template.

    The official GHFDB XLSX template has a non-standard layout:

    - Rows 1-5: Title, description, and metadata (skipped)
    - Row 6:    Technical column headers (used as tablib Dataset headers)
    - Row 7:    Unit labels (skipped)
    - Row 8:    Allowed range of values (skipped)
    - Row 9+:   Data rows

    The data sheet is named ``"data list"``.

    References:
        - Fuchs et al. (2021). A new database structure for the IHFC Global
          Heat Flow Database. Earth System Science Data.
        - Fuchs et al. (2023). The Global Heat Flow Database: Update 2023.
    """

    def create_dataset(self, in_stream: bytes) -> tablib.Dataset:
        """Parse a GHFDB XLSX file and return a tablib Dataset.

        Reads column headers from row 6 and data from row 9 onwards,
        skipping rows 7 (unit labels) and 8 (Allowed range of values) and the metadata rows 1-5.

        Args:
            in_stream: Raw bytes of the uploaded XLSX file.

        Returns:
            tablib.Dataset with headers from row 6 and data from row 9+.
        """
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(in_stream), read_only=True, data_only=True)
        ws = wb["data list"]

        headers = [cell.value for cell in ws[6]]
        dataset = tablib.Dataset(headers=headers)

        for row in ws.iter_rows(min_row=9, values_only=True):
            dataset.append(row)

        wb.close()
        return dataset
