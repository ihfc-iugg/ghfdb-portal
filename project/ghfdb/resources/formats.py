"""
Import format classes for GHFDB import/export resources.

Defines:
- GHFDBImportFormat: Custom XLSX reader for the GHFDB spreadsheet template
"""

from io import BytesIO

import tablib
from import_export.formats.base_formats import XLSX


class GHFDBImportFormat(XLSX):
    """Custom XLSX format for the GHFDB spreadsheet template.

    The official GHFDB XLSX template has a non-standard layout:

    - Rows 1-5: Title, description, and metadata (skipped)
    - Row 6:    Technical column headers (used as tablib Dataset headers)
    - Row 7:    Unit labels (skipped)
    - Row 8:    Allowed range of values (skipped)
    - Row 9+:   Data rows

    The data sheet is named ``"data list"``.

    References:
        - Fuchs et al. (2021). A new database structure for the IHFC Global
          Heat Flow Database. Earth System Science Data.
        - Fuchs et al. (2023). The Global Heat Flow Database: Update 2023.
    """

    def get_title(self) -> str:
        """Human-readable label shown in the admin import format dropdown."""
        return "GHFDB Official Template"

    def create_dataset(self, in_stream: bytes) -> tablib.Dataset:
        """Parse a GHFDB XLSX file and return a tablib Dataset.

        Reads column headers from row 6 and data from row 9 onwards,
        skipping rows 7 (unit labels) and 8 (Allowed range of values) and the metadata rows 1-5.

        Args:
            in_stream: Raw bytes of the uploaded XLSX file.

        Returns:
            tablib.Dataset with headers from row 6 and data from row 9+.
        """
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(in_stream), read_only=True, data_only=True)
        ws = wb["data list"]

        headers = [cell.value for cell in ws[6]]
        dataset = tablib.Dataset(headers=headers)

        for row in ws.iter_rows(min_row=9, values_only=True):
            dataset.append(row)

        wb.close()
        return dataset


class GHFDBSimpleImportFormat(GHFDBImportFormat):
    """Custom XLSX format for GHFDB spreadsheets with a simplified header layout.

    Some submitter templates omit the unit-label and allowed-range rows present
    in the official IHFC template.  The *simple* layout is:

    - Rows 1-5: Arbitrary metadata (skipped)
    - Row 6:    Column headers (identical meaning to the official template)
    - Row 7+:   Data rows (no intervening unit or range rows)

    This class subclasses :class:`GHFDBImportFormat` and overrides only
    ``create_dataset()`` to start data iteration at row 7 instead of row 9.
    All widget, upsert, and field-mapping logic is inherited unchanged.

    References:
        - Fuchs et al. (2021). A new database structure for the IHFC Global
          Heat Flow Database. Earth System Science Data.
        - Fuchs et al. (2023). The Global Heat Flow Database: Update 2023.
    """

    def get_title(self) -> str:
        """Human-readable label shown in the admin import format dropdown."""
        return "GHFDB Simple Template"

    def create_dataset(self, in_stream: bytes) -> tablib.Dataset:
        """Parse a simple-layout GHFDB XLSX file and return a tablib Dataset.

        Reads column headers from row 6 and data from row 7 onwards.
        Rows 1-5 are arbitrary metadata rows that are skipped.  Unlike the
        official template there are no unit-label (row 7) or allowed-range
        (row 8) rows to skip.

        Args:
            in_stream: Raw bytes of the uploaded XLSX file.

        Returns:
            tablib.Dataset with headers from row 6 and data from row 7+.
        """
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(in_stream), read_only=True, data_only=True)
        ws = wb["data list"]

        headers = [cell.value for cell in ws[6]]
        dataset = tablib.Dataset(headers=headers)

        for row in ws.iter_rows(min_row=7, values_only=True):
            dataset.append(row)

        wb.close()
        return dataset
